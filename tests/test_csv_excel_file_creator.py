import pytest
import openpyxl
import csv
from io import BytesIO
from poorlib.filecreator.csv_excel_file_creator import FileCreator  # Assuming the class is in a file named file_creator.py


# Fixture for initializing FileCreator
@pytest.fixture
def file_creator():
    headers = ["Name", "Age", "Occupation"]
    return FileCreator(headers)


def test_initialize_headers(file_creator):
    # Test if headers are correctly initialized in Excel
    assert file_creator.sheet.cell(row=1, column=1).value == "Name"
    assert file_creator.sheet.cell(row=1, column=2).value == "Age"
    assert file_creator.sheet.cell(row=1, column=3).value == "Occupation"

    # Test if headers are stored correctly for CSV
    assert file_creator.rows == []


def test_add_row(file_creator):
    # Add a row and test if it's added correctly to both Excel and CSV data
    row = ["Alice", 30, "Engineer"]
    file_creator.add_row(row)

    # Check if row is added to Excel
    assert file_creator.sheet.cell(row=2, column=1).value == "Alice"
    assert file_creator.sheet.cell(row=2, column=2).value == 30
    assert file_creator.sheet.cell(row=2, column=3).value == "Engineer"

    # Check if row is added to the internal CSV rows
    assert file_creator.rows == [row]


def test_save_excel(file_creator):
    # Add a row and save the Excel file in memory
    file_creator.add_row(["Alice", 30, "Engineer"])
    file_creator.save_excel()

    # Check if the in-memory Excel file is created
    in_memory_file = file_creator.get_file()
    assert isinstance(in_memory_file, BytesIO)

    # Load the Excel file from memory and check content
    in_memory_file.seek(0)
    workbook = openpyxl.load_workbook(in_memory_file)
    sheet = workbook.active

    # Verify headers and row data
    assert sheet.cell(row=1, column=1).value == "Name"
    assert sheet.cell(row=2, column=1).value == "Alice"


def test_save_csv(file_creator):
    # Add a row and save the CSV file in memory
    file_creator.add_row(["Alice", 30, "Engineer"])
    file_creator.save_csv()

    # Check if the in-memory CSV file is created
    in_memory_file = file_creator.get_file()
    assert isinstance(in_memory_file, BytesIO)

    # Load the CSV file from memory and check content
    in_memory_file.seek(0)
    csv_reader = csv.reader(in_memory_file.read().decode().splitlines())
    rows = list(csv_reader)

    # Verify headers and row data
    assert rows[0] == ["Name", "Age", "Occupation"]
    assert rows[1] == ["Alice", "30", "Engineer"]


def test_add_row_validation(file_creator):
    # Test if adding a row with incorrect length raises an error
    with pytest.raises(ValueError, match="Row data must match the number of headers."):
        file_creator.add_row(["Alice", 30])  # Missing one value
